#!/usr/bin/env python3
"""Extract author/title index from Excel into viewer-data/author-index.json."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from constants import (
    AUTHOR_INDEX_FILENAME,
    AUTHOR_INDEX_SHEET,
    AUTHOR_INDEX_XLSX,
    DOCUMENT_TITLE,
    PUBLIC_DATA_DIR,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
AUTHOR_INDEX_PATH = PROJECT_ROOT / PUBLIC_DATA_DIR / AUTHOR_INDEX_FILENAME

CANDIDATE_PATHS = [
    PROJECT_ROOT / "resources" / AUTHOR_INDEX_XLSX,
    PROJECT_ROOT / "resource" / AUTHOR_INDEX_XLSX,
    PROJECT_ROOT / "source" / AUTHOR_INDEX_XLSX,
]

REQUIRED_HEADERS = ("Page", "Names", "Paper Title")


def find_excel_path() -> Path:
    for path in CANDIDATE_PATHS:
        if path.exists():
            return path
    searched = "\n".join(f"  - {p}" for p in CANDIDATE_PATHS)
    raise FileNotFoundError(
        f"Author index Excel not found. Looked for:\n{searched}"
    )


def normalize_text(value: str) -> str:
    """Normalize text for case-insensitive author/title matching."""
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = text.replace(".", " ")
    text = re.sub(r"[^a-z0-9\s\-']", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify_name(name: str) -> str:
    base = normalize_text(name)
    slug = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    return slug or "unknown"


def split_names(raw_names: str) -> list[str]:
    parts = [p.strip() for p in str(raw_names).split(";")]
    return [p for p in parts if p]


def build_aliases(name: str) -> list[str]:
    """Generate searchable aliases for a person name."""
    norm = normalize_text(name)
    if not norm:
        return []

    aliases: list[str] = [norm]
    tokens = norm.split()
    if not tokens:
        return aliases

    first = tokens[0]
    last = tokens[-1]
    middles = tokens[1:-1]

    aliases.append(first)
    aliases.append(last)
    if first != last:
        aliases.append(f"{first} {last}")

    if middles:
        mid_initials = [m[0] for m in middles if m]
        if mid_initials:
            aliases.append(f"{first} {' '.join(mid_initials)} {last}")
            aliases.append(f"{first[0]} {' '.join(mid_initials)} {last}")
            aliases.append(f"{first[0]}{''.join(mid_initials)} {last}")

    if len(first) > 0 and last:
        aliases.append(f"{first[0]} {last}")
        aliases.append(f"{first[0]}{last}")

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique: list[str] = []
    for alias in aliases:
        a = alias.strip()
        if a and a not in seen:
            seen.add(a)
            unique.append(a)
    return unique


def parse_page(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def extract_rows(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{AUTHOR_INDEX_SHEET}' is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(
            f"Missing required headers {missing}. Found: {headers}. "
            f"Expected: {list(REQUIRED_HEADERS)}"
        )

    col = {name: headers.index(name) for name in REQUIRED_HEADERS}
    entries: list[dict] = []
    skipped = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        page = parse_page(row[col["Page"]] if col["Page"] < len(row) else None)
        raw_names = row[col["Names"]] if col["Names"] < len(row) else None
        title = row[col["Paper Title"]] if col["Paper Title"] < len(row) else None

        raw_names_str = str(raw_names).strip() if raw_names is not None else ""
        title_str = str(title).strip() if title is not None else ""
        authors = split_names(raw_names_str)

        if page is None or not title_str or not authors:
            skipped += 1
            print(
                f"WARN: Skipping row {row_idx} "
                f"(page={page!r}, names={raw_names_str!r}, title={title_str!r})"
            )
            continue

        search_parts = [normalize_text(a) for a in authors] + [normalize_text(title_str)]
        for author in authors:
            search_parts.extend(build_aliases(author))
        search_text = " ".join(dict.fromkeys(p for p in search_parts if p))

        entries.append(
            {
                "id": f"paper-{len(entries) + 1:04d}",
                "page": page,
                "title": title_str,
                "authors": authors,
                "rawNames": raw_names_str,
                "searchText": search_text,
            }
        )

    if skipped:
        print(f"Skipped {skipped} incomplete row(s).")
    return entries


def build_authors(entries: list[dict]) -> list[dict]:
    by_key: dict[str, dict] = {}

    for entry in entries:
        for name in entry["authors"]:
            key = normalize_text(name)
            if not key:
                continue
            if key not in by_key:
                by_key[key] = {
                    "id": f"author-{slugify_name(name)}",
                    "name": name,
                    "searchKey": key,
                    "aliases": build_aliases(name),
                    "papers": [],
                }
            paper = {"page": entry["page"], "title": entry["title"]}
            existing = by_key[key]["papers"]
            if not any(p["page"] == paper["page"] and p["title"] == paper["title"] for p in existing):
                existing.append(paper)

    authors = list(by_key.values())
    authors.sort(key=lambda a: a["name"].lower())
    for author in authors:
        author["papers"].sort(key=lambda p: (p["page"], p["title"].lower()))
    return authors


def main() -> int:
    try:
        excel_path = find_excel_path()
    except FileNotFoundError as err:
        print(f"ERROR: {err}", file=sys.stderr)
        return 1

    print(f"Reading: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if AUTHOR_INDEX_SHEET not in wb.sheetnames:
        print(
            f"ERROR: Sheet '{AUTHOR_INDEX_SHEET}' not found. "
            f"Available: {wb.sheetnames}",
            file=sys.stderr,
        )
        wb.close()
        return 1

    try:
        entries = extract_rows(wb[AUTHOR_INDEX_SHEET])
    except ValueError as err:
        print(f"ERROR: {err}", file=sys.stderr)
        wb.close()
        return 1
    finally:
        wb.close()

    authors = build_authors(entries)
    payload = {
        "document": {
            "title": DOCUMENT_TITLE,
            "sourceFile": excel_path.name,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
        },
        "summary": {
            "entryCount": len(entries),
            "authorCount": len(authors),
        },
        "entries": entries,
        "authors": authors,
    }

    AUTHOR_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with AUTHOR_INDEX_PATH.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
        fh.write("\n")

    print(f"Wrote {AUTHOR_INDEX_PATH}")
    print(f"Author index entries: {len(entries)}")
    print(f"Unique authors: {len(authors)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
